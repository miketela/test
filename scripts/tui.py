#!/usr/bin/env python3
"""
Simple terminal interface to run explore/transform/report and pick files.

- Lists available files under `source/`.
- Lets you choose which files to include for explore/transform.
- Creates a temporary source directory with only the selected files
  and runs the pipeline with SOURCE_DIR pointed there.
"""

import os
import re
import shutil
import subprocess
import sys
from pathlib import Path
from typing import List, Tuple, Optional, Set
from collections import Counter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = PROJECT_ROOT / "source"
RAW_DIR = PROJECT_ROOT / "data" / "raw"
TMP_SOURCE_DIR = PROJECT_ROOT / ".tmp_source_run"
TMP_RAW_DIR = PROJECT_ROOT / ".tmp_raw_run"
METRICS_DIR = PROJECT_ROOT / "metrics"

# Persist simple state across actions within the same TUI session
LAST_SELECTED_SUBTYPES: Optional[Set[str]] = None
LAST_RUN_ID: Optional[str] = None


def prompt(msg: str, default: Optional[str] = None) -> str:
    sfx = f" [{default}]" if default else ""
    val = input(f"{msg}{sfx}: ").strip()
    return val or (default or "")

# Optional richer prompts via InquirerPy
try:
    from InquirerPy import inquirer as _inq
    HAS_INQUIRER = True
except Exception:
    HAS_INQUIRER = False


def prompt_checkbox(message: str, choices: List[str], default: Optional[List[str]] = None) -> List[str]:
    if HAS_INQUIRER:
        # Add an explicit "Select All" option for clarity
        all_choice = {'name': '<< Select All >>', 'value': '__ALL__'}
        payload = [all_choice] + [{'name': c, 'value': c} for c in choices]
        # If default equals all choices, preselect the select-all sentinel
        _default = default or []
        if default and set(default) == set(choices):
            _default = ['__ALL__']
        selected = _inq.checkbox(
            message=message,
            choices=payload,
            default=_default,
            transformer=lambda r: f"{len([x for x in r if x != '__ALL__']) if '__ALL__' not in r else len(choices)} selected",
            validate=lambda r: True,
            instruction="Space to toggle, Enter to accept"
        ).execute()
        # Expand select-all sentinel
        if isinstance(selected, list) and '__ALL__' in selected:
            return choices
        return selected
    # Fallback: show indexed list and parse indices
    print(message)
    for i, c in enumerate(choices, 1):
        print(f"  {i}. {c}")
    raw = input("Select by numbers (e.g., 1,3,5-7) or 'a': ").strip()
    idxs = parse_indices(raw, len(choices))
    return [choices[i - 1] for i in idxs]


def prompt_text(message: str, default: Optional[str] = None) -> str:
    if HAS_INQUIRER:
        return _inq.text(message=message + (f" [{default}]" if default else ""), default=default or "").execute().strip()
    return prompt(message, default)

def prompt_confirm(message: str, default: bool = False) -> bool:
    if HAS_INQUIRER:
        return _inq.confirm(message=message, default=default).execute()
    ans = input(f"{message} [{'Y/n' if default else 'y/N'}]: ").strip().lower()
    if not ans:
        return default
    return ans in {"y", "yes", "s", "si", "sí"}


def prompt_select(message: str, choices: List[str], default: Optional[str] = None) -> str:
    if HAS_INQUIRER:
        return _inq.select(message=message, choices=choices, default=default or (choices[0] if choices else None)).execute()
    # Fallback to numeric selection
    print(message)
    for i, c in enumerate(choices, 1):
        print(f"  {i}. {c}")
    idxs = parse_indices(input("Select one: "), len(choices))
    return choices[idxs[0] - 1] if idxs else (default or choices[0])


def prompt_int(message: str, default: Optional[int] = None) -> int:
    default_str = str(default) if default is not None else None
    while True:
        val = prompt_text(message, default_str)
        try:
            return int(val)
        except Exception:
            print("Please enter a valid integer.")


def list_source_files() -> List[Path]:
    if not SOURCE_DIR.exists():
        return []
    files = []
    for p in sorted(SOURCE_DIR.iterdir()):
        if p.is_file() and p.suffix.lower() in {".csv", ".xlsx", ".txt"}:
            files.append(p)
    return files


def print_menu(title: str, options: List[str]) -> int:
    print(f"\n=== {title} ===")
    for i, opt in enumerate(options, 1):
        print(f"  {i}. {opt}")
    while True:
        choice = input("Select an option (number): ").strip()
        if choice.isdigit():
            idx = int(choice)
            if 1 <= idx <= len(options):
                return idx
        print("Invalid selection. Try again.")


def parse_indices(s: str, n: int) -> List[int]:
    # Accept formats: "a" (all), "1,3,5-7"
    s = s.strip().lower()
    if s in {"a", "all", "*"}:
        return list(range(1, n + 1))
    result = set()
    for part in s.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            if a.isdigit() and b.isdigit():
                start, end = int(a), int(b)
                for i in range(start, end + 1):
                    if 1 <= i <= n:
                        result.add(i)
        elif part.isdigit():
            i = int(part)
            if 1 <= i <= n:
                result.add(i)
    return sorted(result)


def pick_files(files: List[Path]) -> List[Path]:
    if not files:
        print("No files found.")
        return []
    names = [p.name for p in files]
    selected_names = prompt_checkbox("Select files (space to toggle):", names, default=names)
    if isinstance(selected_names, list) and any(name == '__ALL__' for name in selected_names):
        selected_names = names
    chosen = [p for p in files if p.name in selected_names]
    print("Selected:")
    for p in chosen:
        print(f"  - {p.name}")
    return chosen


def list_raw_run_files(year: int, month: int) -> List[Path]:
    files: List[Path] = []
    if not RAW_DIR.exists():
        return files
    patt_csv = f"*__run-{year}{month:02d}.csv"
    patt_CSV = f"*__run-{year}{month:02d}.CSV"
    patt_txt = f"*__run-{year}{month:02d}.txt"
    patt_TXT = f"*__run-{year}{month:02d}.TXT"
    # On Windows (case-insensitive FS), the same file can match both patterns.
    # Collect and then deduplicate by resolved path while preserving order.
    candidates: List[Path] = []
    candidates.extend(sorted(RAW_DIR.glob(patt_csv)))
    candidates.extend(sorted(RAW_DIR.glob(patt_CSV)))
    candidates.extend(sorted(RAW_DIR.glob(patt_txt)))
    candidates.extend(sorted(RAW_DIR.glob(patt_TXT)))
    seen = set()
    for p in candidates:
        try:
            rp = p.resolve()
        except Exception:
            rp = p
        if rp not in seen:
            seen.add(rp)
            files.append(p)
    return files

def list_all_raw_files() -> List[Path]:
    """List all raw files across all runs, deduplicated by real path."""
    files: List[Path] = []
    if not RAW_DIR.exists():
        return files
    candidates: List[Path] = []
    candidates.extend(sorted(RAW_DIR.glob("*__run-*.csv")))
    candidates.extend(sorted(RAW_DIR.glob("*__run-*.CSV")))
    candidates.extend(sorted(RAW_DIR.glob("*__run-*.txt")))
    candidates.extend(sorted(RAW_DIR.glob("*__run-*.TXT")))
    seen = set()
    for p in candidates:
        try:
            rp = p.resolve()
        except Exception:
            rp = p
        if rp not in seen:
            seen.add(rp)
            files.append(p)
    return files

def find_latest_run_in_raw() -> Optional[Tuple[int, int]]:
    """Scan data/raw and return the latest (year, month) found in __run-YYYYMM files."""
    if not RAW_DIR.exists():
        return None
    runs: List[Tuple[int, int]] = []
    for p in RAW_DIR.glob("*__run-*.csv"):
        m = re.search(r"__run-(\d{6})", p.name)
        if m:
            yyyymm = m.group(1)
            try:
                runs.append((int(yyyymm[:4]), int(yyyymm[4:6])))
            except Exception:
                continue
    if not runs:
        return None
    runs.sort()
    return runs[-1]


def infer_subtype(filename: str) -> str:
    m = re.match(r"^(.+?)_\d{8}(?:__run-\d+)?\.[A-Za-z0-9]+$", filename)
    if m:
        return m.group(1)
    m = re.search(r"(\d{8})", filename)
    if m:
        return filename[: m.start()].rstrip("_")
    return filename


def prompt_subtype_filter(files: List[Path]) -> Set[str]:
    # Build subtype stats: file counts and approximate row counts
    def _row_count_csv(path: Path) -> int:
        try:
            with path.open('r', encoding='utf-8', errors='ignore') as f:
                c = sum(1 for _ in f)
            return max(c - 1, 0)
        except Exception:
            return 0
    def _row_count_xlsx(path: Path, sheet_sel: Optional[str]) -> int:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            # Resolve sheet by name or index
            ws = None
            if sheet_sel and sheet_sel.isdigit():
                idx = int(sheet_sel)
                if 0 <= idx < len(wb.sheetnames):
                    ws = wb[wb.sheetnames[idx]]
            elif sheet_sel and sheet_sel in wb.sheetnames:
                ws = wb[sheet_sel]
            if ws is None:
                # Fallback to first sheet
                ws = wb[wb.sheetnames[0]]
            # Count non-empty rows (any cell has a value)
            count = 0
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    count += 1
            # Subtract header if present (assume first row is header)
            return max(count - 1, 0)
        except Exception:
            return 0

    file_counts = Counter()
    row_counts = Counter()
    # Ask once for XLSX sheet selection if there are XLSX files
    sheet_sel: Optional[str] = None
    if any(p.suffix.lower() == '.xlsx' for p in files):
        sheet_sel = prompt_text("XLSX sheet (name or 0-based index, blank=first)", "") or None

    for p in files:
        st = infer_subtype(p.name)
        file_counts[st] += 1
        ext = p.suffix.lower()
        if ext == '.csv':
            row_counts[st] += _row_count_csv(p)
        elif ext == '.xlsx':
            row_counts[st] += _row_count_xlsx(p, sheet_sel)

    # Sort subtypes by file count desc, then by name
    subtypes_sorted = sorted(file_counts.keys(), key=lambda k: (-file_counts[k], k))
    if not subtypes_sorted:
        return set()
    print("\nDetected subtypes (files, registros):")
    for i, st in enumerate(subtypes_sorted, 1):
        print(f"  {i}. {st} ({file_counts[st]} files, {row_counts[st]} registros)")
    raw = input("Filter by subtype names (comma-separated), 'a' for all, or leave empty: ").strip()
    if not raw or raw.lower() in {"a", "all", "*"}:
        return set()
    wanted = {s.strip() for s in raw.split(",") if s.strip()}
    wanted_lower = {w.lower() for w in wanted}
    chosen = {st for st in subtypes_sorted if st.lower() in wanted_lower}
    return chosen


def extract_date_from_name(name: str) -> Optional[Tuple[int, int]]:
    # Look for 8-digit date and derive year/month
    m = re.search(r"(\d{8})", name)
    if not m:
        return None
    ymd = m.group(1)
    try:
        return int(ymd[:4]), int(ymd[4:6])
    except Exception:
        return None


def prepare_tmp_source(selected: List[Path]) -> Path:
    if TMP_SOURCE_DIR.exists():
        shutil.rmtree(TMP_SOURCE_DIR)
    TMP_SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    for p in selected:
        shutil.copy2(p, TMP_SOURCE_DIR / p.name)
    return TMP_SOURCE_DIR


def prepare_tmp_raw(selected: List[Path]) -> Path:
    if TMP_RAW_DIR.exists():
        shutil.rmtree(TMP_RAW_DIR)
    TMP_RAW_DIR.mkdir(parents=True, exist_ok=True)
    for p in selected:
        shutil.copy2(p, TMP_RAW_DIR / p.name)
    return TMP_RAW_DIR

def prepare_tmp_raw_with_run(selected: List[Path], year: int, month: int) -> Path:
    """Copy selected files to TMP_RAW_DIR and stamp __run-YYYYMM in filenames."""
    run_id = f"{year}{month:02d}"
    if TMP_RAW_DIR.exists():
        shutil.rmtree(TMP_RAW_DIR)
    TMP_RAW_DIR.mkdir(parents=True, exist_ok=True)
    for p in selected:
        stem = p.stem
        suffix = p.suffix
        # If already has __run, keep it; otherwise, append
        if '__run-' in p.name:
            target_name = p.name
        else:
            target_name = f"{stem}__run-{run_id}{suffix}"
        shutil.copy2(p, TMP_RAW_DIR / target_name)
    return TMP_RAW_DIR


def run_cmd(args: List[str], env_overrides: Optional[dict] = None) -> int:
    env = os.environ.copy()
    if env_overrides:
        env.update(env_overrides)
    print("\n> ", " ".join(args))
    proc = subprocess.run(args, env=env)
    return proc.returncode


def action_explore(selected: List[Path]):
    # Infer period from first file, or ask
    default_period = extract_date_from_name(selected[0].name) if selected else None
    default_year = str(default_period[0]) if default_period else "2024"
    default_month = str(default_period[1]) if default_period else "1"
    year = prompt_int("Year", int(default_year))
    month = prompt_int("Month", int(default_month))

    tmp_dir = prepare_tmp_source(selected)
    print(f"Using temporary SOURCE_DIR: {tmp_dir}")
    rc = run_cmd([sys.executable, str(PROJECT_ROOT / "main.py"),
                  "explore", "--atoms", "AT12", "--year", str(year), "--month", str(month)],
                 env_overrides={"SBP_SOURCE_DIR": str(tmp_dir)})
    if rc == 0:
        print("Explore completed.")
    elif rc == 2:
        print("Explore completed with warnings.")
    else:
        print("Explore failed.")
        # Try to locate the per-run log file and show a short tail
        _show_run_log_tail("explore", f"{year}{month:02d}")
        return

    # Cache last selection for streamlined transform
    global LAST_SELECTED_SUBTYPES, LAST_RUN_ID
    LAST_SELECTED_SUBTYPES = {infer_subtype(p.name) for p in selected}
    LAST_RUN_ID = f"{year}{month:02d}"
    try:
        print(f"Saved selection for next Transform (run {LAST_RUN_ID}): {', '.join(sorted(LAST_SELECTED_SUBTYPES))}")
    except Exception:
        pass

def _show_run_log_tail(command: str, run_id: str, lines: int = 25) -> None:
    """Utility: show tail of the most recent log for a command/run_id."""
    try:
        log_dir = PROJECT_ROOT / "logs" / "AT12" / command
        if not log_dir.exists():
            return
        candidates = [p for p in log_dir.glob(f"*{run_id}*.log") if p.is_file()]
        if not candidates:
            return
        latest = max(candidates, key=lambda p: p.stat().st_mtime)
        print(f"See detailed log: {latest}")
        try:
            content = latest.read_text(encoding="utf-8", errors="ignore").splitlines()
            tail = content[-lines:]
            print(f"Last {lines} log lines:")
            for ln in tail:
                print("  " + ln)
        except Exception:
            pass
    except Exception:
        pass


def _find_and_add_dependencies(chosen: List[Path], year: int, month: int) -> List[Path]:
    """Auto-include AT02/AT03 dependencies if not already chosen by the user."""
    # Avoid modifying the original list
    chosen_with_deps = list(chosen)
    
    # Check which dependencies are already included
    chosen_subtypes = {infer_subtype(p.name) for p in chosen}
    
    # Define required dependencies
    dependencies = {
        "AT02_CUENTAS": f"AT02_CUENTAS_{year}{month:02d}01.csv",
        "AT03_CREDITOS": f"AT03_CREDITOS_{year}{month:02d}01.csv"
    }
    
    # Add missing dependencies
    for subtype, filename in dependencies.items():
        if subtype not in chosen_subtypes:
            dep_path = RAW_DIR / filename
            if dep_path.exists():
                print(f"Found and added required dependency: {filename}")
                chosen_with_deps.append(dep_path)
            else:
                print(f"Warning: Required dependency not found: {filename}")
                
    return chosen_with_deps


def action_transform():
    # New flow: pick files across all runs, then infer period from selection
    global LAST_SELECTED_SUBTYPES, LAST_RUN_ID
    # Fast-path: if we have a recent explore selection, reuse it
    if LAST_RUN_ID and LAST_SELECTED_SUBTYPES:
        try:
            year = int(LAST_RUN_ID[:4])
            month = int(LAST_RUN_ID[4:6])
            available = list_raw_run_files(year, month)
            chosen = [p for p in available if infer_subtype(p.name) in LAST_SELECTED_SUBTYPES]
            if chosen:
                chosen = _find_and_add_dependencies(chosen, year, month)
                tmp_raw = prepare_tmp_raw(chosen)
                try:
                    print(f"Using previous selection (subtypes: {', '.join(sorted(LAST_SELECTED_SUBTYPES))})")
                except Exception:
                    pass
                print(f"Using temporary RAW_DIR: {tmp_raw}")
                rc = run_cmd([sys.executable, str(PROJECT_ROOT / "main.py"),
                              "transform", "--atoms", "AT12", "--year", str(year), "--month", str(month)],
                             env_overrides={"SBP_DATA_RAW_DIR": str(tmp_raw)})
                if rc == 0:
                    print("Transform completed.")
                elif rc == 2:
                    print("Transform completed with warnings.")
                else:
                    print("Transform failed.")
                    _show_run_log_tail("transform", f"{year}{month:02d}")
                return
        except Exception:
            # Fall back to interactive flow
            pass

    all_available = list_all_raw_files()
    if not all_available:
        print("No files found in data/raw (no __run-*.csv files). Fallback to source/...")
        files = list_source_files()
        if not files:
            print("No files in source/ either.")
            return
        wanted = prompt_subtype_filter(files)
        list_for_pick = [p for p in files if not wanted or infer_subtype(p.name) in wanted]
        if not list_for_pick:
            print("No files after applying subtype filter.")
            return
        selected = pick_files(list_for_pick)
        if not selected:
            print("No selection.")
            return
        # Infer or prompt period
        default_period = extract_date_from_name(selected[0].name)
        year = prompt_int("Year", default_period[0] if default_period else 2024)
        month = prompt_int("Month", default_period[1] if default_period else 1)
        selected = _find_and_add_dependencies(selected, year, month)
        tmp_raw = prepare_tmp_raw_with_run(selected, year, month)
        print(f"Using temporary RAW_DIR: {tmp_raw}")
        rc = run_cmd([sys.executable, str(PROJECT_ROOT / "main.py"),
                      "transform", "--atoms", "AT12", "--year", str(year), "--month", str(month)],
                     env_overrides={"SBP_DATA_RAW_DIR": str(tmp_raw)})
        if rc == 0:
            print("Transform completed.")
        elif rc == 2:
            print("Transform completed with warnings.")
        else:
            print("Transform failed.")
            _show_run_log_tail("transform", f"{year}{month:02d}")
        return
    # Optional subtype filter
    wanted = prompt_subtype_filter(all_available)
    list_for_pick = [p for p in all_available if not wanted or infer_subtype(p.name) in wanted]
    if not list_for_pick:
        print("No files after applying subtype filter.")
        return
    chosen = pick_files(list_for_pick)
    if not chosen:
        print("No selection.")
        return

    # Infer run(s) present in the selection
    runs = sorted({re.search(r"__run-(\d{6})", p.name).group(1) for p in chosen if re.search(r"__run-(\d{6})", p.name)})
    selected_run: Optional[str] = None
    if not runs:
        # Fallback to latest run across raw
        latest = find_latest_run_in_raw()
        if not latest:
            print("Could not infer run from selection and no runs detected in data/raw.")
            return
        selected_run = f"{latest[0]}{latest[1]:02d}"
        print(f"No run id in selection; using latest run {selected_run}")
    elif len(runs) == 1:
        selected_run = runs[0]
        print(f"Detected run from selection: {selected_run}")
    else:
        # Multiple runs selected; ask user to pick one
        selected_run = prompt_select("Multiple runs detected. Pick one run (YYYYMM)", runs, default=runs[-1])
        # Filter chosen to only the selected run
        chosen = [p for p in chosen if selected_run in p.name]
        if not chosen:
            print("No files left after narrowing by selected run.")
            return

    year = int(selected_run[:4])
    month = int(selected_run[4:6])

    chosen = _find_and_add_dependencies(chosen, year, month)

    tmp_raw = prepare_tmp_raw(chosen)
    print(f"Using temporary RAW_DIR: {tmp_raw}")
    rc = run_cmd([sys.executable, str(PROJECT_ROOT / "main.py"),
                  "transform", "--atoms", "AT12", "--year", str(year), "--month", str(month)],
                 env_overrides={"SBP_DATA_RAW_DIR": str(tmp_raw)})
    if rc == 0:
        print("Transform completed.")
    elif rc == 2:
        print("Transform completed with warnings.")
    else:
        print("Transform failed.")
        _show_run_log_tail("transform", f"{year}{month:02d}")


def action_report():
    # Pick a metrics json under metrics/
    metric_files = sorted((METRICS_DIR).glob("*.json"))
    if not metric_files:
        print("No metrics JSON files found under metrics/.")
        return
    print("\nAvailable metrics files:")
    for i, p in enumerate(metric_files, 1):
        print(f"  {i}. {p.name}")
    idxs = parse_indices(input("Select one (number): "), len(metric_files))
    if not idxs:
        print("No selection.")
        return
    chosen = metric_files[idxs[0] - 1]
    out = prompt("Output PDF path", str(PROJECT_ROOT / "reports" / "out.pdf"))
    rc = run_cmd([sys.executable, str(PROJECT_ROOT / "main.py"),
                  "report", "--metrics-file", str(chosen), "--output", out])
    if rc != 0:
        print("Report generation failed.")
        # Try to infer run_id from metrics filename and show report log tail
        m = re.search(r"__run-(\d{6})", chosen.stem)
        if m:
            _show_run_log_tail("report", m.group(1))

def _collect_output_files() -> List[Path]:
    """Collect output files for cleanup (testing only)."""
    targets: List[Path] = []
    base = PROJECT_ROOT / "data" / "processed" / "transforms" / "AT12"
    # Incidencias, procesados, consolidated, state
    for sub in ["incidencias", "procesados", "consolidated", "state"]:
        d = base / sub
        if d.exists():
            targets.extend(sorted(d.glob("**/*")))
    # Metrics JSONs
    mdir = PROJECT_ROOT / "metrics"
    if mdir.exists():
        targets.extend(sorted(mdir.glob("*.json")))
    return [p for p in targets if p.is_file()]

def action_clean():
    """Delete output artifacts (testing helper)."""
    print("\n[Testing] Clean outputs: transforms/AT12 + metrics")
    files = _collect_output_files()
    if not files:
        print("Nothing to delete.")
        return
    # Show a short preview
    preview = [str(p.relative_to(PROJECT_ROOT)) for p in files[:10]]
    print("Files to delete (preview):")
    for p in preview:
        print(f"  - {p}")
    if len(files) > 10:
        print(f"  ... and {len(files) - 10} more")
    if not prompt_confirm("Proceed to delete these files?", default=False):
        print("Clean cancelled.")
        return
    removed = 0
    for p in files:
        try:
            p.unlink()
            removed += 1
        except Exception:
            pass
    print(f"Removed {removed} file(s).")


def main():
    print("AT12 Terminal Interface")
    if not HAS_INQUIRER:
        print("(Tip) For interactive menus, install InquirerPy: pip install InquirerPy")
    menu_items = [
        "Explore (pick files)",
        "Transform (pick from raw/source)",
        "Report (choose metrics JSON)",
        "Clean (delete outputs - testing)",  # TODO: remove before release
        "Exit",
    ]
    while True:
        if HAS_INQUIRER:
            selection = prompt_select("Main Menu", menu_items, default=menu_items[0])
        else:
            idx = print_menu("Main Menu", menu_items)
            selection = menu_items[idx - 1]

        if selection.startswith("Explore"):
            files = list_source_files()
            if not files:
                print("No files in source/.")
                continue
            wanted = prompt_subtype_filter(files)
            list_for_pick = [p for p in files if not wanted or infer_subtype(p.name) in wanted]
            selected = pick_files(list_for_pick)
            if selected:
                action_explore(selected)
        elif selection.startswith("Transform"):
            action_transform()
        elif selection.startswith("Report"):
            action_report()
        elif selection.startswith("Clean"):
            action_clean()
        else:
            break


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nAborted.")
